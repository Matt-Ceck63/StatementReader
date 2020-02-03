from openpyxl import *
import pandas as pd


class ExcelSaver:

    def __init__(self, path):
        self.path = path

        self.wb = load_workbook(path)
        last_sheet_name = self.wb.sheetnames[-1]
        self.ws = self.wb[last_sheet_name]

        excel_file = pd.ExcelFile(path)
        self.df = pd.read_excel(excel_file, last_sheet_name)

        pass

    def save(self, values, combo_data, check_data, months):
        # print(values)
        # print(sum(self.to_float(v) for v in values))
        # print(len(values))
        # print(len(combo_data))
        # print(len(check_data))
        # print(len(months))

        for c, cd, m, v in zip(check_data, combo_data, months, values):

            if c is True:

                col = self.get_col_num(m) + 1
                r = self.get_row_num(cd) + 2
                cell = self.to_float(self.ws.cell(row=r, column=col).value)
                cell += self.to_float(v)
                self.ws.cell(row=r, column=col).value = cell

                print("")
                print(m)
                print(cd)
                print(col)
                print(r)
                print(cell)

        self.wb.save(self.path)

    def to_float(self, val):
        try:
            val = float(val)
        except ValueError:
            val = val.replace(",", "")
            try:
                val = float(val)
            except ValueError:
                print("Value Error could not convert string to float: " + val)
        return val

    def get_col_num(self, month):
        columns = list(self.df.columns)

        col = columns.index(month)

        return col

    def get_row_num(self, category):
        # find row index of category
        values = list(self.df["Month"].values)
        index = values.index(category)
        category_index = self.df["Month"].index[index]
        return category_index

    def add_to_cell(self, value, category, month):
        # find row index of category
        values = list(self.df["Month"].values)
        index = values.index(category)
        category_index = self.df["Month"].index[index]

        # go to cell with that index in that column
        cell = self.df[month].values[category_index]
        self.df[month].values[category_index] = cell + value

        return cell + value
